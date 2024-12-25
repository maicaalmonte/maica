import pdfplumber
import openpyxl
import os
from PIL import Image
import pytesseract
import numpy as np  # Import numpy for array handling
import cv2  # Ensure OpenCV is imported

# Define paths
pdf_path = "path_to_your_pdf.pdf"  # Replace with your PDF path
excel_file = "output_excel_file.xlsx"  # Replace with desired output Excel path

# Check if PDF exists
if not os.path.exists(pdf_path):
    print(f"PDF file not found: {pdf_path}")
else:
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    default_sheet.title = "Summary"

    # Open the PDF with pdfplumber
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            print(f"Processing page {page_num}...")

            # Extract tables
            tables = page.extract_tables()
            if tables:
                print(f"Tables found on page {page_num}. Writing to Excel...")
                for table_index, table in enumerate(tables):
                    sheet_name = f"Page{page_num}_Table{table_index + 1}"
                    new_sheet = workbook.create_sheet(title=sheet_name)

                    for row_num, row in enumerate(table, start=1):
                        for col_num, cell in enumerate(row, start=1):
                            new_sheet.cell(row=row_num, column=col_num, value=cell)
            else:
                print(f"No tables found on page {page_num}. Attempting OCR...")

                # Convert page to image
                pil_image = page.to_image(resolution=300).original
                image_array = np.array(pil_image)  # Convert PIL image to numpy array

                # Convert RGB to BGR (for OpenCV)
                bgr_image = cv2.cvtColor(image_array, cv2.COLOR_RGB2BGR)

                # Save the image (optional step)
                temp_image_path = f"page_{page_num}.png"
                cv2.imwrite(temp_image_path, bgr_image)

                # Perform OCR
                text = pytesseract.image_to_string(Image.open(temp_image_path), lang="eng")
                print(f"OCR text from page {page_num}: {text[:200]}...")

                # Write OCR text to Excel
                ocr_sheet_name = f"Page{page_num}_OCR"
                ocr_sheet = workbook.create_sheet(title=ocr_sheet_name)

                for row_num, line in enumerate(text.split("\n"), start=1):
                    ocr_sheet.cell(row=row_num, column=1, value=line)

                # Clean up temporary image
                os.remove(temp_image_path)

    # Remove the default sheet if unused
    if not default_sheet.max_row > 1:
        workbook.remove(default_sheet)

    # Save the Excel file
    workbook.save(excel_file)
    print(f"Excel file saved at: {excel_file}")
